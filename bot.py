from botcity.web import WebBot, Browser, By
from botcity.maestro import *
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.chart import Reference, BarChart3D


# Disable errors if we are not connected to Maestro
BotMaestroSDK.RAISE_NOT_CONNECTED = False


def pesquisar_cidade(bot, cidade):
    bot.browse("https://www.google.com")
    bot.wait(2000)
# https://ws.apicep.com/cep.json?code=69083000
    while len(bot.find_elements('//*[@id="APjFqb"]', By.XPATH)) < 1:
        bot.wait(1000)
        print('carrengado.')
    bot.find_element('//*[@id="APjFqb"]', By.XPATH).send_keys(cidade)
    bot.wait(1000)
    bot.enter()


def extrair_dados_clima(bot):

    count = 0
    dados = []
    while True:
        count += 1
        dia_semana = bot.find_element(
            f'//*[@id="wob_dp"]/div[{count}]/div[1]', By.XPATH).text
        elemento_clicavel = bot.find_element(
            f'//*[@id="wob_dp"]/div[{count}]/div[1]', By.XPATH)
        elemento_clicavel.click()
        max = bot.find_element(
            f'//*[@id="wob_dp"]/div[{count}]/div[3]/div[1]/span[1]', By.XPATH).text
        min = bot.find_element(
            f'//*[@id="wob_dp"]/div[{count}]/div[3]/div[2]/span[1]', By.XPATH).text
        umidade = bot.find_element('//*[@id="wob_hm"]', By.XPATH).text
        umidade = str(umidade).replace("%", "")
        bot.tab()
        bot.enter()
        dados.append(
            {'umidade': umidade, 'dia_semana': dia_semana, 'max': max, 'min': min})
        if count == 8:
            break
    return dados


def salva_dados_temperatura_na_planilha(dados):  # dados é uma lista
    arquivo = 'dados.xlsx'
    wb = openpyxl.load_workbook(arquivo)

    aba_atual = wb.active
    aba_atual[f'A1'] = ' '
    aba_atual[f'B1'] = 'MAX'
    aba_atual[f'C1'] = 'MIN'

    linha = 2
    for dado in dados:
        if linha > 9 : break
        aba_atual[f'A{linha}'] = dado['dia_semana']
        aba_atual[f'B{linha}'] = int(dado['max'])
        aba_atual[f'C{linha}'] = int(dado['min'])
        linha+=1
    wb.save(arquivo)

def salva_dados_umidade_na_planilha(dados):
    arquivo = 'dados.xlsx'
    wb = openpyxl.load_workbook(arquivo)

    aba_atual = wb.active
    aba_atual[f'A20'] = ' '
    aba_atual[f'B20'] = 'Umidade %'

    linha = 21
    for dado in dados:
        if linha > 29 : break
        aba_atual[f'A{linha}'] = dado['dia_semana']
        aba_atual[f'B{linha}'] = int(dado['umidade'])
        linha+=1
    wb.save(arquivo)
    
def gerar_grafico(arquivo_excel, posicao_grafico='E1'):
    # # Carregar o arquivo Excel existente
    wb = openpyxl.load_workbook(arquivo_excel)
    aba_atual = wb.active
    wb.save('dados.xlsx')

    if aba_atual.max_row < 2:
        print("Não há dados suficientes para gerar um gráfico.")
        return
    
    chart = BarChart3D()
    valores = Reference(aba_atual,           
                   min_col=2,  
                   max_col=3, 
                   min_row=1,  
                   max_row=9) 
    dias_semana = Reference(aba_atual, 
                 min_col=1, 
                 max_col=1, 
                 min_row=2, 
                 max_row=9)
    chart.add_data(valores, titles_from_data=True)
    chart.title = "Temperatura Manaus,AM"
    chart.x_axis.title = 'Dias da Semana'
    chart.y_axis.title = 'Temperatura em Graus'
    aba_atual.add_chart(chart,posicao_grafico)
    chart.set_categories(dias_semana)
    wb.save(arquivo_excel)

    chart2 = BarChart3D()

    valores_umi = Reference(aba_atual,           
                   min_col=2,  
                   max_col=3, 
                   min_row=20,  
                   max_row=29) 
    dias_semana = Reference(aba_atual, 
                 min_col=1, 
                 max_col=1, 
                 min_row=20, 
                 max_row=29)
    chart2.title = "Umidade em %"
    chart2.add_data(valores_umi, titles_from_data=True)
    chart2.set_categories(dias_semana)
    aba_atual.add_chart(chart2,'E21')
    wb.save(arquivo_excel)

def main():

    maestro = BotMaestroSDK.from_sys_args()
    execution = maestro.get_execution()

    bot = WebBot()

    bot.headless = False

    bot.browser = Browser.CHROME

    bot.driver_path = ChromeDriverManager().install()

    try:
        arquivo = 'dados.xlsx'
        cidade = "clima Manaus, AM"
        pesquisar_cidade(bot, cidade)
        bot.wait(1000)
        dados=extrair_dados_clima(bot)
        salva_dados_temperatura_na_planilha(dados)
        salva_dados_umidade_na_planilha(dados)
        gerar_grafico(arquivo)

    except Exception as ex:
        print(ex)
        bot.save_screenshot('erro.png')

    finally:
        bot.wait(2000)
        bot.stop_browser()


def not_found(label):
    print(f"Element not found: {label}")


if __name__ == '__main__':
    main()